from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, messagebox
from openpyxl import Workbook, load_workbook
import re
import os
import sys

# Define the path for the Excel file
FILE_PATH = r"C:\Users\91964\OneDrive\Desktop\build\submissions.xlsx"

# Determine if the script is running in a PyInstaller bundle
if getattr(sys, "frozen", False):
    OUTPUT_PATH = Path.cwd()
    ASSETS_PATH = Path(sys._MEIPASS) / "assets/frame0"
else:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / "assets/frame0"

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

def validate_email(email):
    # Regex for validating an email
    pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(pattern, email)

def save_submission():
    customer_name = entry_1.get()
    email_address = entry_2.get()
    feedback = entry_3.get("1.0", "end-1c")
    
    if not customer_name or not email_address or not feedback:
        messagebox.showerror("Error", "All fields must be filled out")
        return
    
    if not validate_email(email_address):
        messagebox.showerror("Error", "Invalid email address")
        return

    if not os.path.exists(FILE_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Submissions"
        sheet.append(["ID", "Customer Name", "Email Address", "Feedback"])
    else:
        workbook = load_workbook(FILE_PATH)
        sheet = workbook.active

    # Get the current max ID
    max_id = sheet.max_row
    
    # Save the new submission with an auto-incremented ID
    sheet.append([max_id, customer_name, email_address, feedback])
    workbook.save(FILE_PATH)

    # Clear the form fields
    entry_1.delete(0, "end")
    entry_2.delete(0, "end")
    entry_3.delete("1.0", "end")

    # Display success message
    messagebox.showinfo("Success", "Feedback submitted successfully")

window = Tk()

window.geometry("600x400")
window.configure(bg = "#FFFFFF")

canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 400,
    width = 600,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    285.0,
    0.0,
    600.0,
    400.0,
    fill="#344865",
    outline="")

canvas.create_rectangle(
    0.0,
    0.0,
    285.0,
    406.0,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    298.0,
    15.0,
    anchor="nw",
    text="Resume Support Feedbacks",
    fill="#FFFFFF",
    font=("Inter ExtraBold", 20 * -1)
)

canvas.create_text(
    300.0,
    52.0,
    anchor="nw",
    text="Customer Name",
    fill="#FFFFFF",
    font=("Inter ExtraBold", 15 * -1)
)

canvas.create_text(
    300.0,
    117.0,
    anchor="nw",
    text="Email Address",
    fill="#FFFFFF",
    font=("Inter ExtraBold", 15 * -1)
)

canvas.create_text(
    298.0,
    191.0,
    anchor="nw",
    text="Feedback",
    fill="#FFFFFF",
    font=("Inter ExtraBold", 15 * -1)
)

entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    438.0,
    94.0,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_1.place(
    x=300.0,
    y=78.0,
    width=276.0,
    height=30.0
)

entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    438.0,
    159.0,
    image=entry_image_2
)
entry_2 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_2.place(
    x=300.0,
    y=143.0,
    width=276.0,
    height=30.0
)

entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    436.0,
    265.5,
    image=entry_image_3
)
entry_3 = Text(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_3.place(
    x=298.0,
    y=217.0,
    width=276.0,
    height=95.0
)

button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=save_submission,  # Connect button to save_submission function
    relief="flat"
)
button_1.place(
    x=338.0,
    y=320.0,
    width=196.0,
    height=61.0
)

button_image_hover_1 = PhotoImage(
    file=relative_to_assets("button_hover_1.png"))

def button_1_hover(e):
    button_1.config(
        image=button_image_hover_1
    )
def button_1_leave(e):
    button_1.config(
        image=button_image_1
    )

button_1.bind('<Enter>', button_1_hover)
button_1.bind('<Leave>', button_1_leave)

image_image_1 = PhotoImage(
    file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(
    142.0,
    203.0,
    image=image_image_1
)
window.resizable(False, False)
window.mainloop()
